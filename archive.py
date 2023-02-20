import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from datetime import date as dt_date
from pandas.tseries.holiday import USFederalHolidayCalendar
from pandas.tseries.offsets import CustomBusinessDay
import tkinter as tk
from tkinter import filedialog

# Create a Tkinter window and hide it
root = tk.Tk()
root.withdraw()

# Open a file dialog window and ask the user to select an Excel file
file_path = filedialog.askopenfilename()

# Load the workbook and select the Vacations sheet
workbook = openpyxl.load_workbook(file_path)
sheet = workbook['Vacations']

# Define a dictionary to store the vacation days and workdays for each employee
result = {}

# Iterate over the rows in the sheet, starting from row 2
for row in range(2, sheet.max_row + 1):
    # Get the employee name, start date, and end date from the columns
    print(row)
    name = sheet.cell(row=row, column=1).value
    start_date = sheet.cell(row=row, column=4).value
    end_date = sheet.cell(row=row, column=5).value

    # Convert the start and end dates to datetime objects
    start_date = datetime.strptime(start_date, '%Y-%m-%d')
    end_date = datetime.strptime(end_date, '%Y-%m-%d')

    # Iterate over all the days in the vacation range, and count the number of vacation days and workdays for each year and month
    curr_date = start_date
    while curr_date <= end_date:
        year = curr_date.year
        month = curr_date.month
        vacation_days = 1
        #cal = USFederalHolidayCalendar()
        #bday = CustomBusinessDay(calendar=cal)
        workdays = len(pd.bdate_range(curr_date, curr_date, freq=bday))
        key = f"{name}-{year}-{month}"
        # If this is the first time we've seen this employee and month, add them to the result dictionary
        if key not in result:
            result[key] = {'vacation_days': vacation_days, 'workdays': workdays}
        # If we've seen this employee and month before, add the new vacation days and workdays to their totals
        else:
            result[key]['vacation_days'] += vacation_days
            result[key]['workdays'] += workdays
        curr_date += timedelta(days=1)

# Print the total number of vacation days and workdays for each employee, separated by year and month
for key, data in result.items():
    name, year, month = key.split('-')
    print(f"{name}: {year}-{month}: Vacation days: {data['vacation_days']}, Workdays: {data['workdays']}")